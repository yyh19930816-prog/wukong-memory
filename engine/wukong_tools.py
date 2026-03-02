# -*- coding: utf-8 -*-
"""
悟空工具箱 — wukong_tools.py
给悟空配备真实可执行的工具，每个工具都返回真实结果。
用于 Function Calling，悟空自主决定用哪个工具。
"""
import os, json, time, requests, subprocess, shutil
from datetime import datetime
from pathlib import Path

# ── 配置 ──────────────────────────────────────────────────────────────────────
FEISHU_APP_ID     = "cli_a92effd632b85cd5"
FEISHU_APP_SECRET = "agCdNI6zfbIjqMBAfF3cmeMzPkPhWFFq"
EVOMAP_URL        = "https://evomap.ai"
EVOMAP_NODE_ID    = "node_wukong_001"
MEMORY_LONG       = r"C:\Users\Administrator\.wukong\workspace\MEMORY.md"
TOOLS_MD          = r"C:\Users\Administrator\.wukong\workspace\TOOLS.md"
WORKSPACE_DIR     = r"C:\Users\Administrator\.wukong\workspace"
DESKTOP           = os.path.join(os.path.expanduser("~"), "Desktop")

# ── 搜索API Key（填入后搜索质量大幅提升）────────────────────────────────────
# Serper.dev：注册免费拿2500次Google搜索 → https://serper.dev
# 注册后在 https://serper.dev/api-key 复制Key填到这里
SERPER_API_KEY    = ""   # 填入后自动启用Google搜索
# Brave Search：注册免费拿2000次/月 → https://api.search.brave.com/app/keys
BRAVE_API_KEY     = ""   # 备选搜索引擎

# ── GitHub API Token（可选，填入后每小时5000次，不填也有60次）──────────────
# 在 https://github.com/settings/tokens 生成 Personal Access Token（不需要任何权限）
GITHUB_TOKEN      = ""   # 可选，不填也能用，只是限速更低

# 允许悟空访问的目录白名单（防止误操作系统文件）
ALLOWED_ROOTS = [
    DESKTOP,
    r"C:\Users\Administrator\Documents",
    r"C:\Users\Administrator\Downloads",
    r"D:\\",
    r"E:\CURSOR",
    WORKSPACE_DIR,
]

# ── 飞书 token 缓存 ────────────────────────────────────────────────────────────
_feishu_token = None
_feishu_token_expire = 0

def _get_feishu_token():
    global _feishu_token, _feishu_token_expire
    if _feishu_token and time.time() < _feishu_token_expire - 60:
        return _feishu_token
    r = requests.post(
        "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": FEISHU_APP_ID, "app_secret": FEISHU_APP_SECRET},
        timeout=10
    )
    d = r.json()
    if d.get("code") == 0:
        _feishu_token = d["tenant_access_token"]
        _feishu_token_expire = time.time() + d.get("expire", 7200)
        return _feishu_token
    return None

# ══════════════════════════════════════════════════════════════════════════════
# 工具实现（每个函数返回字符串结果，悟空读取后用中文回答老板）
# ══════════════════════════════════════════════════════════════════════════════

def tool_get_datetime() -> str:
    """获取当前日期和时间（来自本机系统时钟，这是真实时间，不是EVOMAP或任何API返回的时间戳）"""
    now = datetime.now()
    weekdays = ["星期一","星期二","星期三","星期四","星期五","星期六","星期日"]
    wd = weekdays[now.weekday()]
    result = f"本机系统时间：{now.strftime('%Y年%m月%d日')} {wd} {now.strftime('%H:%M:%S')}"
    result += f"\n（注意：EVOMAP等外部API返回的timestamp字段是服务器记录时间，不代表现在几点）"
    return result


def tool_read_memory(filename: str = "MEMORY.md") -> str:
    """
    读取记忆文件内容
    filename: 文件名，如 MEMORY.md / TOOLS.md / SOUL.md / AGENTS.md / HEARTBEAT.md
              或日期如 2026-03-01（读当日对话记录）
    """
    safe_names = {
        "MEMORY.md": MEMORY_LONG,
        "TOOLS.md": TOOLS_MD,
        "SOUL.md": os.path.join(WORKSPACE_DIR, "SOUL.md"),
        "AGENTS.md": os.path.join(WORKSPACE_DIR, "AGENTS.md"),
        "HEARTBEAT.md": os.path.join(WORKSPACE_DIR, "HEARTBEAT.md"),
    }
    # 日期格式，读当日 memory
    if len(filename) == 10 and filename[4] == '-':
        path = os.path.join(WORKSPACE_DIR, "memory", f"{filename}.md")
    else:
        path = safe_names.get(filename, "")

    if not path or not os.path.exists(path):
        return f"文件不存在：{filename}"
    try:
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        return content[:3000] if len(content) > 3000 else content
    except Exception as e:
        return f"读取失败：{e}"


def tool_write_memory(content: str, section: str = "新增记录") -> str:
    """
    追加内容到 MEMORY.md（老板的偏好/教训/重要信息）
    content: 要写入的内容
    section: 段落标题
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    entry = f"\n\n## [{ts}] {section}\n{content}\n"
    try:
        os.makedirs(os.path.dirname(MEMORY_LONG), exist_ok=True)
        with open(MEMORY_LONG, "a", encoding="utf-8") as f:
            f.write(entry)
        return f"已写入 MEMORY.md — 段落：{section}，时间：{ts}"
    except Exception as e:
        return f"写入失败：{e}"


def tool_query_evomap(action: str = "heartbeat") -> str:
    """
    查询或更新 EVOMAP.AI 网络
    action:
      - heartbeat: 发送心跳，返回当前状态和积分
      - fetch_skills: 拉取推荐技能列表
    """
    ts = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")
    mid = f"wk_{int(time.time())}"

    if action == "heartbeat":
        payload = {
            "protocol": "gep-a2a",
            "protocol_version": "1.0.0",
            "message_type": "heartbeat",
            "message_id": mid,
            "sender_id": EVOMAP_NODE_ID,
            "timestamp": ts,
            "payload": {
                "status": "alive",
                "metrics": {"tasks_completed": 0, "uptime_seconds": int(time.time())}
            }
        }
        try:
            r = requests.post(f"{EVOMAP_URL}/a2a/heartbeat", json=payload, timeout=15)
            data = r.json()
            code = data.get("code", data.get("status", "unknown"))
            credit = data.get("credit_balance", data.get("payload", {}).get("credit_balance", "未知"))
            if r.status_code == 200:
                return f"EVOMAP心跳成功 | HTTP {r.status_code} | code:{code} | 积分:{credit} | 时间:{ts}"
            else:
                return f"EVOMAP心跳返回异常 | HTTP {r.status_code} | 原始响应:{r.text[:300]}"
        except requests.exceptions.ConnectionError:
            return "EVOMAP连接失败：服务器无法访问（ConnectionError）"
        except requests.exceptions.Timeout:
            return "EVOMAP连接超时：15秒无响应"
        except Exception as e:
            return f"EVOMAP请求异常：{e}"

    elif action == "fetch_skills":
        payload = {
            "protocol": "gep-a2a",
            "protocol_version": "1.0.0",
            "message_type": "fetch",
            "message_id": mid,
            "sender_id": EVOMAP_NODE_ID,
            "timestamp": ts,
            "payload": {"filter": {"tags": ["secretary", "assistant", "communication", "content"]}}
        }
        try:
            r = requests.post(f"{EVOMAP_URL}/a2a/fetch", json=payload, timeout=15)
            if r.status_code == 200:
                data = r.json()
                assets = data.get("payload", {}).get("assets", [])
                if not assets:
                    return f"EVOMAP返回空技能列表 | HTTP 200 | 原始:{r.text[:200]}"
                lines = [f"EVOMAP技能列表（共{len(assets)}个）："]
                for a in assets[:8]:
                    lines.append(f"- {a.get('name','未命名')} | {a.get('description','')[:60]}")
                return "\n".join(lines)
            elif r.status_code == 502:
                return "EVOMAP技能接口故障（HTTP 502 Bad Gateway）— 这是服务器问题，不是悟空的问题。心跳接口正常，但技能学习暂时不可用。积分当前为0，即使接口恢复也可能需要积分才能获取技能。"
            else:
                return f"EVOMAP拉取技能失败 | HTTP {r.status_code} | {r.text[:300]}"
        except requests.exceptions.ConnectionError:
            return "EVOMAP连接失败：服务器无法访问（ConnectionError）"
        except requests.exceptions.Timeout:
            return "EVOMAP连接超时"
        except Exception as e:
            return f"EVOMAP请求异常：{e}"

    return f"未知action：{action}，支持的action：heartbeat / fetch_skills"


def tool_send_feishu(message: str, chat_id: str = "") -> str:
    """
    通过飞书机器人发送消息
    message: 消息内容
    chat_id: 群或用户的chat_id（空则发到默认群）
    """
    token = _get_feishu_token()
    if not token:
        return "飞书token获取失败，请检查App ID和Secret"

    # 读取保存的默认chat_id
    default_chat_file = os.path.join(WORKSPACE_DIR, ".feishu_default_chat")
    if not chat_id and os.path.exists(default_chat_file):
        try:
            with open(default_chat_file, "r") as f:
                chat_id = f.read().strip()
        except:
            pass

    if not chat_id:
        return "没有指定飞书chat_id，也没有默认群。请先告诉我发给谁。"

    try:
        r = requests.post(
            "https://open.feishu.cn/open-apis/im/v1/messages",
            params={"receive_id_type": "chat_id"},
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={
                "receive_id": chat_id,
                "msg_type": "text",
                "content": json.dumps({"text": message})
            },
            timeout=15
        )
        d = r.json()
        if d.get("code") == 0:
            msg_id = d.get("data", {}).get("message_id", "")
            return f"飞书消息发送成功 | code:0 | message_id:{msg_id}"
        else:
            return f"飞书发送失败 | code:{d.get('code')} | msg:{d.get('msg')} | 原始:{r.text[:200]}"
    except Exception as e:
        return f"飞书发送异常：{e}"


def _search_serper(query: str, num: int = 6) -> list:
    """Serper.dev Google搜索"""
    if not SERPER_API_KEY:
        return []
    try:
        r = requests.post(
            "https://google.serper.dev/search",
            headers={"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"},
            json={"q": query, "num": num, "gl": "cn", "hl": "zh-cn"},
            timeout=12
        )
        if r.status_code == 200:
            data = r.json()
            results = []
            if data.get("answerBox"):
                ab = data["answerBox"]
                results.append({"title": ab.get("title", "直接答案"),
                                 "snippet": ab.get("answer", ab.get("snippet", "")),
                                 "url": ab.get("link", ""), "source": "Google知识卡片"})
            for item in data.get("organic", [])[:num]:
                results.append({"title": item.get("title", ""),
                                 "snippet": item.get("snippet", ""),
                                 "url": item.get("link", ""), "source": "Google/Serper"})
            return results
    except:
        pass
    return []


def _search_brave(query: str, num: int = 6) -> list:
    """Brave Search API"""
    if not BRAVE_API_KEY:
        return []
    try:
        r = requests.get(
            "https://api.search.brave.com/res/v1/web/search",
            headers={"Accept": "application/json", "Accept-Encoding": "gzip",
                     "X-Subscription-Token": BRAVE_API_KEY},
            params={"q": query, "count": num, "search_lang": "zh"},
            timeout=12
        )
        if r.status_code == 200:
            data = r.json()
            return [{"title": i.get("title", ""), "snippet": i.get("description", ""),
                     "url": i.get("url", ""), "source": "Brave"}
                    for i in data.get("web", {}).get("results", [])[:num]]
    except:
        pass
    return []


def _search_sogou(query: str, num: int = 5) -> list:
    """搜狗搜索抓取（中文效果好，带Session绕过反爬）"""
    import re
    results = []
    try:
        session = requests.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0",
            "Accept-Language": "zh-CN,zh;q=0.9",
            "Referer": "https://www.sogou.com/",
        })
        # 先拿cookie
        session.get("https://www.sogou.com/", timeout=6)
        r = session.get("https://www.sogou.com/web", params={"query": query, "ie": "utf8"}, timeout=12)
        if r.status_code == 200 and len(r.text) > 10000:
            clean = lambda s: re.sub(r'<[^>]+>', '', s).strip()
            titles = re.findall(r'<h3[^>]*>(.*?)</h3>', r.text, re.DOTALL)
            # 搜狗摘要class随版本变化，用通用方案
            snippets = re.findall(r'class="[^"]*str[^"]*"[^>]*>(.*?)</(?:p|div|span)>', r.text, re.DOTALL)
            if not snippets:
                snippets = re.findall(r'<p[^"]*>((?:[^<]|<[^/]){20,})</p>', r.text, re.DOTALL)
            urls = re.findall(r'<a[^>]*href="(https?://[^"]{10,})"[^>]*class="[^"]*result[^"]*"', r.text)
            if not urls:
                urls = re.findall(r'<h3[^>]*>.*?<a[^>]*href="([^"]+)"', r.text, re.DOTALL)
            added = 0
            for i, title in enumerate(titles):
                t = clean(title)
                if not t or len(t) < 4 or added >= num:
                    continue
                s = clean(snippets[i]) if i < len(snippets) else ""
                u = urls[i] if i < len(urls) else ""
                if "sogou" in u.lower() and "link" in u.lower():
                    u = ""
                results.append({"title": t, "snippet": s[:200], "url": u, "source": "搜狗"})
                added += 1
    except:
        pass
    return results


def _search_360(query: str, num: int = 5) -> list:
    """360搜索 JSON API（稳定，有真实摘要）"""
    import re
    results = []
    try:
        # 360搜索有个JSON接口
        r = requests.get(
            "https://www.so.com/s",
            params={"q": query, "src": "www_normal", "f": "7", "pn": "1"},
            timeout=12,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/122.0.0.0",
                "Accept": "text/html,application/xhtml+xml",
                "Accept-Language": "zh-CN,zh;q=0.9",
            }
        )
        if r.status_code == 200:
            text = r.text
            clean = lambda s: re.sub(r'<[^>]+>', '', s).strip()
            # 360的结果嵌在JSON数据里
            json_results = re.findall(r'"title"\s*:\s*"([^"]+)"[^}]*"description"\s*:\s*"([^"]*)"[^}]*"url"\s*:\s*"([^"]+)"', text)
            if json_results:
                for title, desc, url in json_results[:num]:
                    if title:
                        results.append({"title": title, "snippet": desc[:200], "url": url, "source": "360搜索"})
                return results
            # 备用HTML抓取
            # 360使用React，内容在__INITIAL_STATE__或data-props里
            state_match = re.search(r'window\.__INITIAL_STATE__\s*=\s*({.*?});', text, re.DOTALL)
            if state_match:
                try:
                    import json as jsonlib
                    state = jsonlib.loads(state_match.group(1))
                    items = state.get("search", {}).get("items", []) or state.get("data", {}).get("result", [])
                    for item in items[:num]:
                        t = item.get("title", "") or item.get("name", "")
                        s = item.get("description", "") or item.get("summary", "")
                        u = item.get("url", "") or item.get("link", "")
                        if t:
                            results.append({"title": clean(t), "snippet": clean(s)[:200], "url": u, "source": "360搜索"})
                except:
                    pass
    except:
        pass
    return results


def _search_google_news_rss(query: str, num: int = 6) -> list:
    """Google新闻RSS（完全免费，无需Key，真实新闻结果）"""
    import re
    results = []
    try:
        r = requests.get(
            "https://news.google.com/rss/search",
            params={"q": query, "hl": "zh-CN", "gl": "CN", "ceid": "CN:zh-Hans"},
            timeout=12,
            headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
        )
        if r.status_code == 200:
            titles = re.findall(r"<title>(.*?)</title>", r.text)[1:]
            links = re.findall(r"<link>(.*?)</link>", r.text)[1:]
            sources = re.findall(r"<source[^>]*>(.*?)</source>", r.text)
            dates = re.findall(r"<pubDate>(.*?)</pubDate>", r.text)
            for i, title in enumerate(titles[:num]):
                if not title or title == "Google 新闻":
                    continue
                url = links[i] if i < len(links) else ""
                src = sources[i] if i < len(sources) else ""
                date = dates[i][:16] if i < len(dates) else ""
                snippet = f"来源：{src}" if src else ""
                if date:
                    snippet += f"  日期：{date}"
                results.append({
                    "title": title,
                    "snippet": snippet,
                    "url": url,
                    "source": "Google新闻"
                })
    except:
        pass
    return results


def _search_baidu_mip(query: str, num: int = 5) -> list:
    """百度移动版搜索（绕过PC验证）"""
    import re
    results = []
    try:
        r = requests.get(
            "https://m.baidu.com/s",
            params={"word": query, "from": "1000539b"},
            timeout=12,
            headers={
                "User-Agent": "Mozilla/5.0 (Linux; Android 11; Pixel 5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.91 Mobile Safari/537.36",
                "Accept-Language": "zh-CN,zh;q=0.9",
            }
        )
        if r.status_code == 200 and len(r.text) > 5000:
            clean = lambda s: re.sub(r'<[^>]+>', '', s).strip()
            text = r.text
            # 百度移动版title
            titles = re.findall(r'class="[^"]*c-title[^"]*"[^>]*>(.*?)</(?:a|div|h3)>', text, re.DOTALL)
            snippets = re.findall(r'class="[^"]*c-abstract[^"]*"[^>]*>(.*?)</(?:span|div|p)>', text, re.DOTALL)
            urls = re.findall(r'<a[^>]*href="(https?://[^"]{15,})"[^>]*class="[^"]*c-title[^"]*"', text)
            added = 0
            for i, title in enumerate(titles):
                t = clean(title)
                if not t or len(t) < 4 or added >= num:
                    continue
                s = clean(snippets[i]) if i < len(snippets) else ""
                u = urls[i] if i < len(urls) else ""
                results.append({"title": t, "snippet": s[:200], "url": u, "source": "百度"})
                added += 1
    except:
        pass
    return results


def _format_results(query: str, results: list) -> str:
    """格式化搜索结果为可读文本"""
    engine = results[0]["source"] if results else "未知"
    lines = [f"搜索「{query}」— 来源：{engine}，共{len(results)}条\n"]
    for i, res in enumerate(results, 1):
        t = res.get("title", "")
        s = res.get("snippet", "")
        u = res.get("url", "")
        if not t:
            continue
        lines.append(f"{i}. {t}")
        if s:
            lines.append(f"   {s[:200]}")
        if u:
            lines.append(f"   {u}")
        lines.append("")
    return "\n".join(lines)


def tool_search_web(query: str) -> str:
    """
    搜索网络获取真实信息，返回带来源的搜索结果
    query: 搜索关键词（中英文均可）
    """
    # 优先级：Serper(Google) > Brave > Google新闻RSS（无需Key）> 百度移动版 > 搜狗 > 360
    results = (
        _search_serper(query)
        or _search_brave(query)
        or _search_google_news_rss(query)
        or _search_baidu_mip(query)
        or _search_sogou(query)
        or _search_360(query)
    )

    # 过滤掉无意义的结果
    junk_titles = {"大家还在搜", "实时智能回复", "", "Google 新闻"}
    results = [r for r in results
               if r.get("title", "") not in junk_titles
               and len(r.get("title", "")) > 5]

    if not results:
        return f"搜索失败：搜索「{query}」无结果，网络可能不通。"

    return _format_results(query, results)


def tool_deep_research(topic: str, questions: str = "", include_github: str = "auto") -> str:
    """
    深度研究一个话题：自动联合 Google + GitHub，多角度搜索，整合成有来源的研究报告。
    topic: 研究主题，如"抖音博主张三的内容风格"或"飞书机器人开发"
    questions: 可选，指定要回答的具体问题，用逗号分隔
    include_github: 'auto'=自动判断(技术类话题加入GitHub) / 'yes'=强制搜GitHub / 'no'=不搜GitHub
    """
    import re

    # 判断是否为技术类话题（自动决定是否搜GitHub）
    tech_keywords = ["代码", "开发", "api", "sdk", "爬虫", "机器人", "bot", "python",
                     "编程", "工具", "框架", "库", "插件", "脚本", "接口", "集成",
                     "github", "开源", "部署", "服务器", "自动化", "程序", "软件",
                     "系统", "飞书", "微信", "抖音", "数据"]
    topic_lower = topic.lower()
    is_tech = include_github == "yes" or (
        include_github == "auto" and
        any(kw in topic_lower for kw in tech_keywords)
    )

    # 拆解搜索角度
    if questions:
        search_queries = [q.strip() for q in questions.split(",") if q.strip()]
    else:
        search_queries = [
            topic,
            topic + " 最新 2026",
            topic + " 方法 技巧",
        ]

    junk_titles = {"大家还在搜", "实时智能回复", ""}
    all_results = []
    searched = []
    sections = []

    # ── 第一部分：网络搜索（Google/搜狗/360）──────────────────────────────
    web_results = []
    for q in search_queries[:3]:
        res = (_search_serper(q, 5) or _search_brave(q, 5)
               or _search_google_news_rss(q, 5)
               or _search_baidu_mip(q, 5) or _search_sogou(q, 5) or _search_360(q, 5))
        if res:
            res = [r for r in res
                   if r.get("title", "") not in junk_titles
                   and len(r.get("title", "")) > 5]
        if res:
            searched.append(q)
            for r in res[:3]:
                snippet = r.get("snippet", "")
                url = r.get("url", "")
                entry = f"• {r['title']}"
                if snippet:
                    entry += f"\n  {snippet[:180]}"
                if url:
                    entry += f"\n  来源: {url[:80]}"
                web_results.append(entry)

    if web_results:
        sections.append("【网络资讯】\n" + "\n\n".join(web_results))

    # ── 第二部分：GitHub搜索（技术类话题）─────────────────────────────────
    github_results = []
    if is_tech:
        github_query = topic
        headers = {"User-Agent": "WukongBot/1.0", "Accept": "application/vnd.github+json"}
        if GITHUB_TOKEN:
            headers["Authorization"] = f"token {GITHUB_TOKEN}"
        try:
            r = requests.get(
                "https://api.github.com/search/repositories",
                params={"q": github_query, "sort": "stars", "order": "desc", "per_page": 6},
                headers=headers, timeout=15
            )
            if r.status_code == 200:
                items = r.json().get("items", [])
                total = r.json().get("total_count", 0)
                for item in items[:5]:
                    desc = item.get("description", "") or "无描述"
                    entry = (f"• {item['full_name']}  ★{item['stargazers_count']}"
                             f"  [{item.get('language','?')}]\n"
                             f"  {desc[:100]}\n"
                             f"  {item['html_url']}")
                    github_results.append(entry)
                if github_results:
                    sections.append(
                        f"【GitHub开源方案】（共{total}个相关仓库）\n" +
                        "\n\n".join(github_results)
                    )
        except:
            pass

    # ── 汇总 ─────────────────────────────────────────────────────────────
    if not sections:
        return f"深度研究失败：「{topic}」在所有渠道均无结果，网络可能不通。"

    report = f"深度研究报告：{topic}\n"
    if searched:
        report += f"搜索角度：{' | '.join(searched)}\n"
    if is_tech and github_results:
        report += f"GitHub找到 {len(github_results)} 个开源方案\n"
    report += "=" * 50 + "\n\n"
    report += "\n\n".join(sections)
    return report


def tool_open_url(url: str) -> str:
    """
    打开任意网页链接，读取页面的主要文字内容。
    支持：普通网页、微博、知乎、公众号、新闻等文字类页面。
    注意：抖音/微信视频号等视频类平台只能读到文字信息，无法看视频本身。
    url: 完整的网址，如 https://www.douyin.com/user/xxx
    """
    if not url.startswith("http"):
        url = "https://" + url

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate",
        "Connection": "keep-alive",
    }

    import re

    # 特殊平台提示
    special_tips = {
        "douyin.com": "抖音",
        "tiktok.com": "TikTok",
        "bilibili.com": "B站",
        "youtube.com": "YouTube",
        "weixin.qq.com": "微信公众号",
    }
    platform = next((v for k, v in special_tips.items() if k in url), None)

    try:
        r = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
        r.encoding = r.apparent_encoding or "utf-8"
        html = r.text

        # 提取 title
        title_match = re.search(r"<title[^>]*>(.*?)</title>", html, re.IGNORECASE | re.DOTALL)
        title = re.sub(r"<[^>]+>", "", title_match.group(1)).strip() if title_match else ""

        # 提取 meta description
        desc_match = re.search(r'<meta[^>]+name=["\']description["\'][^>]+content=["\'](.*?)["\']', html, re.IGNORECASE)
        desc = desc_match.group(1).strip() if desc_match else ""

        # 提取 og:description（社交媒体常用）
        og_desc = re.search(r'<meta[^>]+property=["\']og:description["\'][^>]+content=["\'](.*?)["\']', html, re.IGNORECASE)
        og_desc = og_desc.group(1).strip() if og_desc else ""

        # 提取正文（去掉脚本/样式，取纯文字）
        clean_html = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL | re.IGNORECASE)
        clean_html = re.sub(r"<style[^>]*>.*?</style>", "", clean_html, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", clean_html)
        text = re.sub(r"\s+", " ", text).strip()

        lines = [f"网页内容（{url}）："]
        if title:
            lines.append(f"标题：{title}")
        if desc:
            lines.append(f"描述：{desc[:200]}")
        elif og_desc:
            lines.append(f"描述：{og_desc[:200]}")

        # 取正文前800字
        if text and len(text) > 50:
            lines.append(f"正文内容：{text[:800]}")

        if platform in ("抖音", "TikTok", "B站", "YouTube"):
            lines.append(f"（注意：{platform}是视频平台，这里只能读到文字信息，无法观看视频内容。如需了解视频内容，请告诉我博主名字，我去搜索相关介绍）")

        if len(lines) <= 1:
            return f"页面内容为空或需要登录才能访问：{url}"

        return "\n".join(lines)

    except requests.exceptions.Timeout:
        return f"访问超时（15秒）：{url}\n可能需要登录或该页面需要特殊网络"
    except requests.exceptions.ConnectionError:
        return f"无法访问该链接：{url}\n可能需要VPN或该网址有误"
    except Exception as e:
        return f"打开链接失败：{e}\n链接：{url}"


def tool_write_capability(name: str, trigger: str, input_desc: str,
                          output_desc: str, tool_chain: str,
                          failure_points: str, status: str = "待验证") -> str:
    """
    把一个可复用的解法抽象成能力轮廓，写入 CAPABILITIES.md。
    当发现"这个套路下次还能用"时调用。
    name: 能力名称，如"抖音博主分析"
    trigger: 触发场景描述
    input_desc: 输入是什么
    output_desc: 输出是什么
    tool_chain: 用了哪些工具，顺序是什么
    failure_points: 哪里容易失败，怎么处理
    status: 候选/待验证/已内生化
    """
    # 兼容两个可能的路径
    cap_file = os.path.join(WORKSPACE_DIR, "CAPABILITIES.md")
    if not os.path.exists(cap_file):
        # 尝试相对于本文件的路径
        alt_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "workspace", "CAPABILITIES.md")
        if os.path.exists(alt_path):
            cap_file = alt_path
        else:
            # 自动创建基础文件
            try:
                os.makedirs(os.path.dirname(cap_file), exist_ok=True)
                with open(cap_file, "w", encoding="utf-8") as f:
                    f.write("# 悟空 · 能力轮廓库（CAPABILITIES）\n\n## 🟢 已内生化能力\n\n## 🟡 待验证能力（已抽象，未完整测试）\n\n## 🔵 待抽象池（发现了套路，还没正式写轮廓）\n\n## 📊 进化统计\n- 最近一次更新：" + datetime.now().strftime("%Y-%m-%d") + "\n")
            except Exception as e:
                return f"CAPABILITIES.md 不存在且无法创建：{e}"

    now = datetime.now().strftime("%Y-%m-%d")
    # 生成能力ID（基于时间戳）
    cap_id = f"CAP-{datetime.now().strftime('%m%d%H%M')}"

    entry = f"""
### {cap_id}：{name}
- 触发场景：{trigger}
- 输入：{input_desc}
- 输出：{output_desc}
- 工具链：{tool_chain}
- 失败点：{failure_points}
- 状态：{status}
- 抽象时间：{now}
"""
    try:
        with open(cap_file, "r", encoding="utf-8") as f:
            content = f.read()

        # 根据状态插入到对应区域
        if status == "已内生化":
            insert_after = "## 🟢 已内生化能力"
        elif status == "待验证":
            insert_after = "## 🟡 待验证能力（已抽象，未完整测试）"
        else:
            insert_after = "## 🔵 待抽象池（发现了套路，还没正式写轮廓）"

        if insert_after in content:
            idx = content.index(insert_after) + len(insert_after)
            content = content[:idx] + "\n" + entry + content[idx:]
        else:
            content += "\n" + entry

        # 更新统计时间
        content = content.replace(
            content[content.rfind("- 最近一次更新："):content.rfind("- 最近一次更新：")+30],
            f"- 最近一次更新：{now}"
        )

        with open(cap_file, "w", encoding="utf-8") as f:
            f.write(content)

        return f"能力轮廓已写入：{cap_id}「{name}」→ 状态：{status}"
    except Exception as e:
        return f"写入能力轮廓失败：{e}"


def tool_search_github(query: str, search_type: str = "repo") -> str:
    """
    搜索GitHub获取代码库、技术方案、工具
    query: 搜索关键词，如 'feishu bot python' 或 '抖音爬虫'
    search_type: 'repo'=搜仓库(默认) / 'topic'=搜话题标签
    """
    import base64
    headers = {
        "User-Agent": "WukongBot/1.0",
        "Accept": "application/vnd.github+json",
    }
    if GITHUB_TOKEN:
        headers["Authorization"] = f"token {GITHUB_TOKEN}"

    try:
        if search_type == "topic":
            r = requests.get(
                "https://api.github.com/search/topics",
                params={"q": query, "per_page": 5},
                headers={**headers, "Accept": "application/vnd.github.mercy-preview+json"},
                timeout=15
            )
        else:
            r = requests.get(
                "https://api.github.com/search/repositories",
                params={"q": query, "sort": "stars", "order": "desc", "per_page": 8},
                headers=headers,
                timeout=15
            )

        if r.status_code == 200:
            data = r.json()
            items = data.get("items", [])
            total = data.get("total_count", 0)

            if not items:
                return f"GitHub搜索「{query}」：没有找到相关仓库"

            lines = [f"GitHub搜索「{query}」— 共{total}个结果，显示前{len(items)}个\n"]
            for i, item in enumerate(items, 1):
                name = item.get("full_name", "")
                desc = item.get("description", "") or "无描述"
                stars = item.get("stargazers_count", 0)
                url = item.get("html_url", "")
                lang = item.get("language", "") or "未知"
                updated = item.get("updated_at", "")[:10]
                lines.append(f"{i}. {name}  ★{stars}  [{lang}]  更新:{updated}")
                lines.append(f"   {desc[:100]}")
                lines.append(f"   {url}")
                lines.append("")
            return "\n".join(lines)

        elif r.status_code == 403:
            rate = r.headers.get("X-RateLimit-Remaining", "?")
            reset = r.headers.get("X-RateLimit-Reset", "")
            return f"GitHub API限速（剩余次数:{rate}），请稍后再试或填入GITHUB_TOKEN提升限额"
        else:
            return f"GitHub搜索失败 HTTP {r.status_code}"

    except Exception as e:
        return f"GitHub搜索异常：{e}"


def tool_read_github_repo(repo: str, filepath: str = "README.md") -> str:
    """
    读取GitHub仓库里的文件内容（默认读README）
    repo: 仓库路径，如 'larksuite/oapi-sdk-python'
    filepath: 文件路径，如 'README.md' 或 'examples/send_message.py'
    """
    import base64
    headers = {
        "User-Agent": "WukongBot/1.0",
        "Accept": "application/vnd.github+json",
    }
    if GITHUB_TOKEN:
        headers["Authorization"] = f"token {GITHUB_TOKEN}"

    try:
        # 先获取文件内容
        url = f"https://api.github.com/repos/{repo}/contents/{filepath}"
        r = requests.get(url, headers=headers, timeout=15)

        if r.status_code == 200:
            data = r.json()
            if isinstance(data, list):
                # 是目录，列出文件
                lines = [f"仓库 {repo}/{filepath} 是个目录，包含：\n"]
                for f_item in data[:20]:
                    ftype = "📁" if f_item["type"] == "dir" else "📄"
                    lines.append(f"  {ftype} {f_item['name']}")
                return "\n".join(lines)

            # 是文件，解码内容
            encoding = data.get("encoding", "")
            content_raw = data.get("content", "")
            if encoding == "base64":
                content = base64.b64decode(content_raw).decode("utf-8", errors="replace")
            else:
                content = content_raw

            # 限制长度，太长的文件截取前3000字
            size = data.get("size", 0)
            if len(content) > 3000:
                content = content[:3000] + f"\n\n[文件共{size}字节，已截取前3000字符]"

            return f"文件：{repo}/{filepath}\n大小：{size}字节\n\n{content}"

        elif r.status_code == 404:
            # 尝试自动找README
            if filepath == "README.md":
                for alt in ["readme.md", "README.rst", "README", "docs/README.md"]:
                    r2 = requests.get(
                        f"https://api.github.com/repos/{repo}/contents/{alt}",
                        headers=headers, timeout=10
                    )
                    if r2.status_code == 200:
                        d2 = r2.json()
                        raw = base64.b64decode(d2.get("content", "")).decode("utf-8", errors="replace")
                        return f"文件：{repo}/{alt}\n\n{raw[:3000]}"
            return f"文件不存在：{repo}/{filepath}\n请检查路径是否正确"
        elif r.status_code == 403:
            return "GitHub API限速，请稍后再试"
        else:
            return f"读取失败 HTTP {r.status_code}"

    except Exception as e:
        return f"读取GitHub文件异常：{e}"


def tool_list_files(directory: str = "workspace") -> str:
    """
    列出目录下的文件
    directory: workspace / memory / engine / hud
    """
    base = r"C:\Users\Administrator\.wukong"
    dir_map = {
        "workspace": os.path.join(base, "workspace"),
        "memory": os.path.join(base, "workspace", "memory"),
        "engine": r"E:\CURSOR\wukong-memory\engine",
        "hud": r"E:\CURSOR\wukong-memory\hud",
    }
    path = dir_map.get(directory, os.path.join(base, directory))
    if not os.path.exists(path):
        return f"目录不存在：{directory}"
    try:
        files = os.listdir(path)
        if not files:
            return f"{directory}/ 目录为空"
        return f"{directory}/ 目录文件列表：\n" + "\n".join(f"- {f}" for f in sorted(files))
    except Exception as e:
        return f"列目录失败：{e}"


def _is_allowed(path: str) -> bool:
    """检查路径是否在允许范围内"""
    p = os.path.abspath(path)
    for root in ALLOWED_ROOTS:
        try:
            if p.startswith(os.path.abspath(root)):
                return True
        except:
            pass
    return False


def tool_browse_desktop(subpath: str = "") -> str:
    """
    浏览桌面或桌面下的子目录，列出所有文件和文件夹（含大小、修改时间）
    subpath: 可选，桌面下的子目录路径，如 "文档文件" 或空字符串表示桌面根目录
    """
    target = os.path.join(DESKTOP, subpath) if subpath else DESKTOP
    target = os.path.normpath(target)

    if not _is_allowed(target):
        return f"不允许访问该路径：{target}"
    if not os.path.exists(target):
        return f"路径不存在：{target}"
    if not os.path.isdir(target):
        return f"这是文件不是目录，用 read_file 工具来读它的内容"

    try:
        items = os.listdir(target)
        if not items:
            return f"目录为空：{target}"
        lines = [f"[桌面] {target}（共 {len(items)} 项）："]
        dirs, files = [], []
        for name in sorted(items):
            full = os.path.join(target, name)
            try:
                mtime = datetime.fromtimestamp(os.path.getmtime(full)).strftime("%Y-%m-%d %H:%M")
                if os.path.isdir(full):
                    dirs.append(f"  [文件夹] {name}/  [{mtime}]")
                else:
                    size = os.path.getsize(full)
                    sz = f"{size/1024:.0f}KB" if size > 1024 else f"{size}B"
                    files.append(f"  [文件] {name}  {sz}  [{mtime}]")
            except:
                dirs.append(f"  ? {name}")
        lines.extend(dirs)
        lines.extend(files)
        return "\n".join(lines)
    except Exception as e:
        return f"浏览失败：{e}"


def tool_read_file(path: str) -> str:
    """
    读取电脑上任意文件的内容（支持 txt/md/py/json/csv/log 等文本格式）
    path: 文件完整路径，如 C:\\Users\\Administrator\\Desktop\\合同.txt
          或桌面相对路径，如 desktop:\\合同.txt
    """
    # 支持 desktop:\ 简写
    if path.lower().startswith("desktop:\\") or path.lower().startswith("desktop:/"):
        path = os.path.join(DESKTOP, path[9:])
    path = os.path.normpath(path)

    if not _is_allowed(path):
        return f"不允许访问该路径（不在许可目录内）：{path}"
    if not os.path.exists(path):
        return f"文件不存在：{path}"
    if os.path.isdir(path):
        return f"这是目录不是文件，用 browse_desktop 工具来浏览目录"

    ext = os.path.splitext(path)[1].lower()
    # 二进制格式提示
    binary_exts = {'.exe','.dll','.pyd','.so','.zip','.rar','.7z','.png','.jpg','.jpeg','.gif','.mp4','.mp3'}
    if ext in binary_exts:
        size = os.path.getsize(path)
        return f"这是二进制文件（{ext}），无法直接读取内容。文件大小：{size/1024:.0f}KB，路径：{path}"

    # Word/Excel 文件提示
    if ext in ('.docx', '.doc'):
        try:
            import docx
            doc = docx.Document(path)
            text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return f"Word文档内容（{path}）：\n{text[:3000]}" + ("…（内容过长已截断）" if len(text) > 3000 else "")
        except ImportError:
            return f"读取Word文件需要安装 python-docx 库。文件路径：{path}\n运行：pip install python-docx"
        except Exception as e:
            return f"Word文件读取失败：{e}"

    if ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames[:3]:
                ws = wb[sheet]
                lines.append(f"[Sheet: {sheet}]")
                for row in list(ws.iter_rows(values_only=True))[:50]:
                    row_str = "\t".join(str(c) if c is not None else "" for c in row)
                    if row_str.strip():
                        lines.append(row_str)
            return f"Excel内容（{path}）：\n" + "\n".join(lines[:200])
        except ImportError:
            return f"读取Excel文件需要安装 openpyxl 库。文件路径：{path}\n运行：pip install openpyxl"
        except Exception as e:
            return f"Excel文件读取失败：{e}"

    # 普通文本文件
    encodings = ["utf-8", "gbk", "utf-8-sig", "gb18030"]
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc) as f:
                content = f.read()
            if len(content) > 4000:
                return f"文件内容（{path}，前4000字）：\n{content[:4000]}\n…（文件共{len(content)}字，已截断）"
            return f"文件内容（{path}）：\n{content}"
        except UnicodeDecodeError:
            continue
        except Exception as e:
            return f"读取失败：{e}"
    return f"文件编码无法识别：{path}"


def tool_search_files(keyword: str, directory: str = "desktop", search_content: bool = False) -> str:
    """
    在指定目录搜索文件名或文件内容
    keyword: 搜索关键词
    directory: 搜索范围，desktop=桌面，documents=文档，downloads=下载，d=D盘，e=E盘
    search_content: True=同时搜索文件内容（较慢），False=只搜索文件名
    """
    dir_map = {
        "desktop":   DESKTOP,
        "documents": os.path.join(os.path.expanduser("~"), "Documents"),
        "downloads": os.path.join(os.path.expanduser("~"), "Downloads"),
        "d":         "D:\\",
        "e":         "E:\\CURSOR",
        "workspace": WORKSPACE_DIR,
    }
    search_root = dir_map.get(directory.lower(), DESKTOP)
    if not os.path.exists(search_root):
        return f"搜索目录不存在：{search_root}"

    matches = []
    kw_lower = keyword.lower()

    try:
        for dirpath, dirnames, filenames in os.walk(search_root):
            # 跳过隐藏目录和系统目录
            dirnames[:] = [d for d in dirnames if not d.startswith('.') and d not in ('__pycache__', 'node_modules', '$RECYCLE.BIN')]
            for fname in filenames:
                full = os.path.join(dirpath, fname)
                # 文件名匹配
                if kw_lower in fname.lower():
                    mtime = datetime.fromtimestamp(os.path.getmtime(full)).strftime("%Y-%m-%d")
                    size = os.path.getsize(full)
                    matches.append(f"[文件名匹配] {full}  {size/1024:.0f}KB  {mtime}")
                    if len(matches) >= 20:
                        break
                # 内容匹配（仅文本文件）
                elif search_content and os.path.getsize(full) < 500*1024:
                    ext = os.path.splitext(fname)[1].lower()
                    if ext in ('.txt','.md','.py','.json','.csv','.log','.ini','.cfg'):
                        try:
                            for enc in ["utf-8","gbk"]:
                                try:
                                    with open(full, "r", encoding=enc) as f:
                                        text = f.read()
                                    if kw_lower in text.lower():
                                        idx = text.lower().find(kw_lower)
                                        snippet = text[max(0,idx-30):idx+60].replace('\n',' ')
                                        matches.append(f"[内容匹配] {full}\n    …{snippet}…")
                                    break
                                except UnicodeDecodeError:
                                    continue
                        except:
                            pass
            if len(matches) >= 20:
                break

        if not matches:
            scope = "文件名+内容" if search_content else "文件名"
            return f"在 {search_root} 中搜索 [{scope}] 关键词「{keyword}」，没有找到任何匹配文件"

        result = f"搜索「{keyword}」在 {search_root}，找到 {len(matches)} 个：\n"
        result += "\n".join(matches)
        return result

    except Exception as e:
        return f"搜索失败：{e}"


# ══════════════════════════════════════════════════════════════════════════════
# Function Calling 工具定义（OpenAI格式，DeepSeek-V3支持）
# ══════════════════════════════════════════════════════════════════════════════
TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "get_datetime",
            "description": "获取当前日期和时间",
            "parameters": {"type": "object", "properties": {}, "required": []}
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_memory",
            "description": "读取悟空的记忆文件，包括老板偏好、历史对话、技能手册等",
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string",
                        "description": "文件名：MEMORY.md / TOOLS.md / SOUL.md / AGENTS.md / HEARTBEAT.md，或日期如2026-03-01"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_memory",
            "description": "把重要信息/教训/老板偏好写入长期记忆（MEMORY.md）",
            "parameters": {
                "type": "object",
                "properties": {
                    "content": {"type": "string", "description": "要写入的内容"},
                    "section": {"type": "string", "description": "段落标题，如：老板偏好 / 教训 / 客户信息"}
                },
                "required": ["content"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "query_evomap",
            "description": "查询或更新EVOMAP.AI进化网络，发送心跳或拉取技能",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": ["heartbeat", "fetch_skills"],
                        "description": "heartbeat=发心跳检查状态和积分，fetch_skills=拉取推荐技能列表"
                    }
                },
                "required": ["action"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "send_feishu",
            "description": "通过飞书机器人发送消息给老板或群",
            "parameters": {
                "type": "object",
                "properties": {
                    "message": {"type": "string", "description": "消息内容"},
                    "chat_id": {"type": "string", "description": "飞书chat_id，不知道就留空"}
                },
                "required": ["message"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_web",
            "description": "搜索网络获取实时信息",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "搜索关键词"}
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "列出悟空自己的记忆/工具目录下的文件",
            "parameters": {
                "type": "object",
                "properties": {
                    "directory": {
                        "type": "string",
                        "enum": ["workspace", "memory", "engine", "hud"],
                        "description": "要查看的目录"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "open_url",
            "description": "打开任意网页链接，读取页面内容。老板发来一个链接想让悟空看看是什么，就用这个工具。支持新闻、微博、知乎、公众号等。抖音/B站只能读文字信息。",
            "parameters": {
                "type": "object",
                "properties": {
                    "url": {
                        "type": "string",
                        "description": "要打开的完整网址，如 https://www.douyin.com/user/xxx"
                    }
                },
                "required": ["url"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "browse_desktop",
            "description": "浏览老板电脑的桌面或桌面下的子目录，列出所有文件和文件夹。想知道桌面有什么文件就用这个。",
            "parameters": {
                "type": "object",
                "properties": {
                    "subpath": {
                        "type": "string",
                        "description": "桌面下的子目录名，如'文档文件'或'项目'，空字符串表示桌面根目录"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "读取老板电脑上的文件内容，支持txt/md/Word/Excel/json/csv/py等格式。老板说'帮我看看某个文件'就用这个。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "文件的完整路径，如 C:\\Users\\Administrator\\Desktop\\合同.txt，或桌面简写 desktop:\\合同.txt"
                    }
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": "在老板电脑的指定目录里搜索文件。老板说'帮我找一下某某文件'就用这个。",
            "parameters": {
                "type": "object",
                "properties": {
                    "keyword": {
                        "type": "string",
                        "description": "要搜索的关键词，可以是文件名的一部分"
                    },
                    "directory": {
                        "type": "string",
                        "enum": ["desktop", "documents", "downloads", "d", "e", "workspace"],
                        "description": "搜索范围：desktop=桌面，documents=文档，downloads=下载，d=D盘，e=E盘"
                    },
                    "search_content": {
                        "type": "boolean",
                        "description": "是否同时搜索文件内容（True=搜文件名+内容，False=只搜文件名，默认False）"
                    }
                },
                "required": ["keyword"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "deep_research",
            "description": "深度研究一个话题：自动联合Google+GitHub，多角度搜索，整合成有来源的研究报告。老板说'帮我研究/分析/查资料'时用这个，比search_web全面10倍。技术类话题会自动搜GitHub开源方案。",
            "parameters": {
                "type": "object",
                "properties": {
                    "topic": {
                        "type": "string",
                        "description": "研究主题，如'抖音博主涨粉策略'、'飞书机器人开发'、'2026年短视频行业趋势'"
                    },
                    "questions": {
                        "type": "string",
                        "description": "可选：指定要回答的具体问题，用逗号分隔，如'他发什么内容,粉丝多少,涨粉速度'"
                    },
                    "include_github": {
                        "type": "string",
                        "enum": ["auto", "yes", "no"],
                        "description": "是否搜GitHub：auto=自动判断(技术话题自动搜)，yes=强制搜，no=不搜"
                    }
                },
                "required": ["topic"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_capability",
            "description": "把一个可复用的解法抽象成能力轮廓写入CAPABILITIES.md。当完成一个有价值的任务后，发现'这个套路下次还能用'时调用。这是悟空自我进化的核心动作。",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "能力名称，简洁，如'抖音博主竞品分析'"},
                    "trigger": {"type": "string", "description": "什么情况下触发这个能力"},
                    "input_desc": {"type": "string", "description": "需要哪些输入信息"},
                    "output_desc": {"type": "string", "description": "产出什么结果"},
                    "tool_chain": {"type": "string", "description": "用了哪些工具，顺序是什么"},
                    "failure_points": {"type": "string", "description": "哪里容易失败，如何处理"},
                    "status": {
                        "type": "string",
                        "enum": ["候选", "待验证", "已内生化"],
                        "description": "能力状态：候选=刚发现套路，待验证=已抽象待测试，已内生化=稳定可用"
                    }
                },
                "required": ["name", "trigger", "input_desc", "output_desc", "tool_chain", "failure_points"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_github",
            "description": "在GitHub上搜索开源代码库和技术方案。老板问'有没有现成的XXX工具/代码/框架'，或者遇到技术问题想找解决方案时用这个。GitHub上有全世界最好的开源代码。",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "搜索关键词，如'feishu bot python'、'抖音爬虫'、'wechat auto reply'"
                    },
                    "search_type": {
                        "type": "string",
                        "enum": ["repo", "topic"],
                        "description": "repo=搜仓库(默认)，topic=搜话题标签"
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_github_repo",
            "description": "读取GitHub仓库里的文件内容，默认读README。用于了解某个开源项目怎么用、有哪些功能。",
            "parameters": {
                "type": "object",
                "properties": {
                    "repo": {
                        "type": "string",
                        "description": "仓库路径，如'larksuite/oapi-sdk-python'或'microsoft/autogen'"
                    },
                    "filepath": {
                        "type": "string",
                        "description": "文件路径，默认'README.md'，也可以是'docs/quickstart.md'或具体代码文件"
                    }
                },
                "required": ["repo"]
            }
        }
    }
]

# ══════════════════════════════════════════════════════════════════════════════
# 工具调度器（根据函数名调用对应工具，返回结果字符串）
# ══════════════════════════════════════════════════════════════════════════════
TOOL_DISPATCH = {
    "get_datetime":   lambda args: tool_get_datetime(),
    "read_memory":    lambda args: tool_read_memory(args.get("filename", "MEMORY.md")),
    "write_memory":   lambda args: tool_write_memory(args.get("content", ""), args.get("section", "新增记录")),
    "query_evomap":   lambda args: tool_query_evomap(args.get("action", "heartbeat")),
    "send_feishu":    lambda args: tool_send_feishu(args.get("message", ""), args.get("chat_id", "")),
    "search_web":     lambda args: tool_search_web(args.get("query", "")),
    "deep_research":  lambda args: tool_deep_research(
                         args.get("topic", ""),
                         args.get("questions", ""),
                         args.get("include_github", "auto")
                      ),
    "write_capability": lambda args: tool_write_capability(
                         args.get("name", ""),
                         args.get("trigger", ""),
                         args.get("input_desc", ""),
                         args.get("output_desc", ""),
                         args.get("tool_chain", ""),
                         args.get("failure_points", ""),
                         args.get("status", "待验证")
                      ),
    "search_github":  lambda args: tool_search_github(
                         args.get("query", ""),
                         args.get("search_type", "repo")
                      ),
    "read_github_repo": lambda args: tool_read_github_repo(
                         args.get("repo", ""),
                         args.get("filepath", "README.md")
                      ),
    "list_files":     lambda args: tool_list_files(args.get("directory", "workspace")),
    "open_url":       lambda args: tool_open_url(args.get("url", "")),
    "browse_desktop": lambda args: tool_browse_desktop(args.get("subpath", "")),
    "read_file":      lambda args: tool_read_file(args.get("path", "")),
    "search_files":   lambda args: tool_search_files(
                         args.get("keyword", ""),
                         args.get("directory", "desktop"),
                         args.get("search_content", False)
                      ),
}

def execute_tool(name: str, arguments: dict) -> str:
    """调度并执行工具，返回结果字符串"""
    fn = TOOL_DISPATCH.get(name)
    if not fn:
        return f"未知工具：{name}"
    try:
        return fn(arguments)
    except Exception as e:
        return f"工具执行出错 [{name}]：{e}"
